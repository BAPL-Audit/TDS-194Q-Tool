from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

THRESHOLD = 50_00_000
TDS_RATE = 0.001


def normalize(text):
    if pd.isna(text):
        return None
    return (
        str(text)
        .upper()
        .strip()
        .replace("&", "AND")
        .replace(".", "")
        .replace(",", "")
    )


def _coerce_numeric(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str).str.replace(",", "", regex=False).str.strip(),
        errors="coerce",
    ).fillna(0)


def read_purchase(df):
    df.columns = [str(c).lower().strip() for c in df.columns]

    party_cols = [
        "party name",
        "vendor name",
        "supplier name",
        "ledger name",
        "account name",
        "particulars",
    ]
    amount_cols = [
        "taxable value",
        "amount",
        "net amount",
        "purchase amount",
        "debit",
        "credit",
    ]

    norm_cols = {c.replace(" ", ""): c for c in df.columns}

    party_col = next(
        (norm_cols[c.replace(" ", "")] for c in party_cols if c.replace(" ", "") in norm_cols),
        None,
    )
    amount_col = next(
        (norm_cols[c.replace(" ", "")] for c in amount_cols if c.replace(" ", "") in norm_cols),
        None,
    )

    if not party_col:
        raise ValueError(f"❌ Party column not found. Available columns: {list(df.columns)}")
    if not amount_col:
        raise ValueError(f"❌ Amount column not found. Available columns: {list(df.columns)}")

    df = df.rename(columns={party_col: "party", amount_col: "amount"})
    df["party_norm"] = df["party"].apply(normalize)
    df["amount"] = _coerce_numeric(df["amount"])

    return df[["party", "party_norm", "amount"]]


def read_books(df):
    import re

    df.columns = [str(c).lower().strip() for c in df.columns]

    party_candidates = ["particulars", "party name", "ledger name", "account name"]
    party_col = next((c for c in party_candidates if c in df.columns), None)
    if not party_col:
        raise ValueError(
            f"❌ Party column not found in Books. Available columns: {list(df.columns)}"
        )

    pan_candidates = [
        "pan",
        "pan no",
        "pan no.",
        "pan number",
        "pan of party",
        "pan of deductee",
    ]
    pan_col = next((c for c in pan_candidates if c in df.columns), None)
    if not pan_col:
        raise ValueError(
            f"❌ PAN column not found in Books. Available columns: {list(df.columns)}"
        )

    tds_keywords = [
        "194q",
        "tds194q",
        "tds (194q)",
        "tds-194q",
        "purchase tds",
        "tds on purchase",
        "tds",
    ]

    tds_col = None
    for col in df.columns:
        col_clean = re.sub(r"[^a-z0-9]", "", col)
        for kw in tds_keywords:
            kw_clean = re.sub(r"[^a-z0-9]", "", kw)
            if kw_clean in col_clean:
                tds_col = col
                break
        if tds_col:
            break

    if not tds_col:
        raise ValueError(
            f"❌ TDS (194Q) column not detected in Books. Available columns: {list(df.columns)}"
        )

    df = df.rename(columns={party_col: "party", pan_col: "pan", tds_col: "tds_books"})
    df["party_norm"] = df["party"].apply(normalize)
    df["pan_norm"] = df["pan"].astype(str).str.upper().str.strip()
    df["tds_books"] = _coerce_numeric(df["tds_books"])

    return df[["party", "party_norm", "pan_norm", "tds_books"]]


def read_26q(df):
    df.columns = [str(c).lower().strip() for c in df.columns]

    party_cols = ["name of the deductee", "name of deductee", "deductee name"]
    pan_cols = ["pan of the deductee", "pan of deductee", "last pan of the dedutee"]
    amount_cols = ["amount paid credited rs.", "amount paid", "gross amount"]
    tds_cols = ["tds rs.", "tds amount"]

    party_col = next((c for c in party_cols if c in df.columns), None)
    if not party_col:
        raise ValueError(
            f"❌ Party column not found in 26Q. Available columns: {list(df.columns)}"
        )
    df = df.rename(columns={party_col: "party"})

    pan_col = next((c for c in pan_cols if c in df.columns), None)
    if pan_col:
        df = df.rename(columns={pan_col: "pan"})
    else:
        df["pan"] = ""

    amount_col = next((c for c in amount_cols if c in df.columns), None)
    if amount_col:
        df = df.rename(columns={amount_col: "gross_26q"})
    else:
        df["gross_26q"] = 0

    tds_col = next((c for c in tds_cols if c in df.columns), None)
    if tds_col:
        df = df.rename(columns={tds_col: "tds_26q"})
    else:
        df["tds_26q"] = 0

    df["party_norm"] = df["party"].apply(normalize)
    df["pan_norm"] = df["pan"].astype(str).str.upper().str.strip()
    df["gross_26q"] = _coerce_numeric(df["gross_26q"])
    df["tds_26q"] = _coerce_numeric(df["tds_26q"])

    return df[["party", "party_norm", "pan_norm", "gross_26q", "tds_26q"]]


def compute_194q(purchase):
    summary = purchase.groupby("party_norm", as_index=False).agg(total_purchase=("amount", "sum"))

    summary["liable_amount"] = summary["total_purchase"].apply(lambda x: max(0, x - THRESHOLD))
    summary["tds_required"] = (summary["liable_amount"] * TDS_RATE).round(0)

    return summary


def export_to_excel_openpyxl(dfs: dict):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    output.seek(0)

    wb = load_workbook(output)

    header_font = Font(name="Century Gothic", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Century Gothic", size=10, bold=False, color="000000")
    header_fill = PatternFill(fill_type="solid", start_color="1F4E79", end_color="1F4E79")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            ws.column_dimensions[col_letter].width = max_length + 2

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    return final_output
